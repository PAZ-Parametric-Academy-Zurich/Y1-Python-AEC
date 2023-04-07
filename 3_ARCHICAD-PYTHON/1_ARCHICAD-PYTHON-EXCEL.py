"""
>>> PAZ_Academy = [["PAZ - Parametric Academy Zurich", "https://pazacademy.ch", "info@pazacademy.ch", "+41 44 243 62 16"],["PAZ Central" , "Niederdorfstrasse 77", "CH-8001 Zurich"]]

>>>  Y1 - Python AEC
        Python AEC is a full day Python course for Architects, Engineers and BIM professionals.
        Analyze, Optimize and Automatize with Python your normal day AEC workflows !
        

        
                     .::::::::::::::::::::::::::::::::::::::::::::::::::::::.                                                                                                                           
                  -#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@%+.                                                                                                                       
                +@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@%@@@@@@#-                                                                                                                     
              .%@@@@*:                                                       -#@@@@+                                                                                                                    
              =@@@@-                                                           #@@@@.                                                                                                                   
              +@@@@.                                                           +@@@@.                                                                                                                   
              +@@@@.                                                           +@@@@.                                                                                                                   
              +@@@@.                                                           +@@@@.                                         .............               .......            ...................        
              +@@@@.                   :*+.             -*+                    +@@@@.                                         +@@@@@@@@@@@@@@%*-          @@@@@@@=           %@@@@@@@@@@@@@@@@@@.       
              +@@@@.                  =@@@%.           .@@@%.                  +@@@@.                                         +@@@@@@%%%%%@@@@@@#.       +@@@@@@@%.          #@@@@@@@@@@@@@@@@@=        
              +@@@@.                 :@@@@@-           =@@@@%.                 +@@@@.                                         +@@@@%.      =@@@@@@.     :@@@@@@@@@%                    =@@@@@#.         
              +@@@@.                .@@@@@@=           *@@@@@*                 +@@@@.                                         +@@@@%.       +@@@@@=    .@@@@+.%@@@@+                 .*@@@@@#           
           .-=%@@@@.                +@@@@@@-           +@@@@@@-                +@@@@*=:.                                      +@@@@%       :%@@@@@-    *@@@#  -@@@@@.               -%@@@@%=            
        .*@@@@@@@@@.                %@@@@@@:           -@@@@@@%                +@@@@@@@@%-                                    +@@@@@#****#%@@@@@@=    *@@@#    *@@@@#              #@@@@@#.             
       =@@@@@@@@@@@.               :@@@@@@%.           .@@@@@@@.               +@@@@@@@@@@%:                                  +@@@@@@@@@@@@@@@@*.    -@@@@:    .@@@@@+           .#@@@@@=               
      *@@@@@@@@@@@@.               =@@@@@@=             *@@@@@@:               +@@@@@@@@@@@@=                                 +@@@@@+======-:.       %@@@@%#%%%%@@@@@@-         -@@@@@@=                
     =@@@@@@@@@@@@@.               =@@@@@%              :@@@@@@-               +@@@@@@@@@@@@@-                                +@@@@%                =@@@@@@@@@@@@@@@@@%        +@@@@@%.                 
    .@@@@@@@@@@@@@@.               -@@@@@:               -@@@@@:               +@@@@@@@@@@@@@*                                +@@@@%.              :@@@@#--------=@@@@@#     .%@@@@@+.                  
    .@@@@@@@@@@@@@@.                #@@%.                 -@@@#                +@@@@@@@@@@@@@#                                +@@@@%.             :@@@@%          -@@@@@*   =@@@@@@@@@@@@@@@@@@%        
     #@@@@@@@@@@@@@.                .++                    :+=                 +@@@@@@@@@@@@@=                                +@@@@%              #@@@@-           *@@@@@. .@@@@@@@@@@@@@@@@@@@%        
     .%@@@@@@@@@@@@.            -+-.                           :++:            +@@@@@@@@@@@@%.                                 .....              ...:.             ....:   ....................        
      .%@@@@@@@@@@@.           -@@@@%*=:                  .-+#@@@@%            +@@@@@@@@@@@*                                                                                                            
        -#@@@@@@@@@.           .@@@@@@@@@@%%#***+++***#%@@@@@@@@@@#            +@@@@@@@@@*:                                                                                                             
           -+#@@@@@.            =@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@-            +@@@@%#+-                                                                                                                
              +@@@@.             *@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@+             +@@@@.                                                                                                                   
              +@@@@.              %@@@@@@@@@@@@@@@@@@@@@@@@@@@@@+              +@@@@.                                                                                                                   
              +@@@@.              .*@@@@@@@@@@@@@@@@@@@@@@@@@@@+               +@@@@.                                            .:       :-:       :     .:::.     .:::::.  :.     ::  :.    .:        
              +@@@@.                .+%@@@@@@@@@@@@@@@@@@@@@%=.                +@@@@.                                           .#*%.   *#-.:+#:   ##%.   =@=:-+#=  *%-:::. .@%-   =@@- :#*  -%-        
              +@@@@.                   -*%@@@@@@@@@@@@@@@%+.                   +@@@@.                                           *- *%. =@:        +* +%.  =@:   -@: *@---   :%+@: .#:@-   +%+*          
              +@@@@.                      .-+##%%%%%#*=-.                      +@@@@.                                          -@==+@+ -@:       -@+=+@+  =@:   -@: *%-::   :% *# *:.@-    +@           
              +@@@@.                                                           +@@@@:                                         :@=...=@. *%:  -%:.@=...-@- =@- .=@+  *%.     :%  #@# .@-    =%           
              +@@@@:                                                           *@@@@.                                         -=     --  :+++=. -=     -+ :++++-.   -+++++=  =  :+.  +:    :=           
              :@@@@%-                                                        .=@@@@#                                                                                                                    
                #@@@@@%#####################################################@@@@@@*                                                                                                                     
                 -#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@*:                                                                                                                      
                   .=+*#####################################################*+-
                   

                   
>>>  1_ARCHICAD-PYTHON-EXCEL.py
        Let's start installing Archicad-Python > First install Python enviroment > https://www.python.org/downloads/
        Then we need to activate Python inside Archicad > https://gsdownloads-cdn.azureedge.net/cdn/Python/ARCHICAD-Python_Connection_Getting_Started.pdf (page 1)
        Next if we wanna use Python IDLE with Archicad we need to install with cmd > python –m pip install archicad  
        Remember to have the proper versions Archicad & pip same Archicad 25 > pip install archicad==25.3000 < Release history > https://pypi.org/project/archicad/
        And now let's enjoy with Archicad Python :D !
"""
#This is an starting Archicad Python example from > https://graphisoft.com/downloads/python < Excel File Exporter
from archicad import ACConnection, handle_dependencies #Let's conect our IDLE with a Running Archicad & Bring some Extra Python libraries ;)
from typing import List #We will have modules coming from the Standard Python Library, here a nice one to control Types
import os, sys #os operative system module & systmem helps to control and generate computer files ;)

handle_dependencies('openpyxl') #Here is the name of the external library 'openpyxl' > https://openpyxl.readthedocs.io/en/stable/ ;)

from openpyxl import Workbook #If you don't have this library use cmd (command) > pip install openpyxl < https://pypi.org/project/openpyxl/

conn = ACConnection.connect() #Let's conect our IDLE with a Running Archicad :D !
assert conn

acc = conn.commands
act = conn.types
acu = conn.utilities

scriptFolder = os.path.dirname(os.path.realpath(__file__)) #Location to save our Excel File ;)

################################ CONFIGURATION #################################
worksheetTitlesAndElements = { #Let's Create a Dictionary ;)
    "Beams": acc.GetElementsByType("Beam"), #Here the elments we wanna export information to Excel ;)
    "Walls": acc.GetElementsByType("Wall"), #We will create One Excel Page for Each Item
    "ALL": acc.GetAllElements() #As we learned in our Hello Archicad Python we can get all our Elements ;)
}
propertyUserIds = [ #Now we will work with a List to get the Properties of our Elements ;)
    act.BuiltInPropertyUserId("General_ElementID"), #This Properties will be presented inside each Excel Column
    act.BuiltInPropertyUserId("General_Height"), 
    act.BuiltInPropertyUserId("General_Width"),
    act.BuiltInPropertyUserId("General_Thickness"),
    act.BuiltInPropertyUserId("General_Area"),
    act.BuiltInPropertyUserId("General_OwnerID")
]
outputFolder = scriptFolder
outputFileName = "BeamAndWallGeometry.xlsx" #Name of our Future Excel File ;)
################################################################################

#From this line we will work on the Genration of the Excel File with openpyxl > https://openpyxl.readthedocs.io/en/stable/index.html
def AutoFitWorksheetColumns(ws):
    for columnCells in ws.columns:
        length = max(len(str(cell.value)) for cell in columnCells)
        ws.column_dimensions[columnCells[0].column_letter].width = length


def PrintWorksheetContent(ws):
    for columnCells in ws.columns:
        for cell in columnCells:
            print(f"{ws.title}!{cell.column_letter}{cell.row}={cell.value}")


def FillExcelWorksheetWithPropertyValuesOfElements(ws, propertyIds: List[act.PropertyIdArrayItem], elements: List[act.ElementIdArrayItem]):
    propertyValuesDictionary = acu.GetPropertyValuesDictionary(elements, propertyIds)
    propertyDefinitionsDictionary = dict(zip(propertyIds, acc.GetDetailsOfProperties(propertyIds)))

    ws.cell(row=2, column=1).value = "Element Guid"
    row = 3
    for element, valuesDictionary in propertyValuesDictionary.items():
        ws.cell(row=row, column=1).value = str(element.elementId.guid)
        column = 2
        for propertyId, propertyValue in valuesDictionary.items():
            if row == 3:
                ws.cell(row=1, column=column).value = str(propertyId.propertyId.guid)
                propertyDefinition = propertyDefinitionsDictionary[propertyId].propertyDefinition
                ws.cell(row=2, column=column).value = f"{propertyDefinition.group.name} / {propertyDefinition.name}"
            ws.cell(row=row, column=column).value = propertyValue
            column += 1
        row += 1

    AutoFitWorksheetColumns(ws)
    PrintWorksheetContent(ws)


propertyIds = acc.GetPropertyIds(propertyUserIds)
wb = Workbook()
ws = wb.active

i = 0
for title, elements in worksheetTitlesAndElements.items():
    if i == 0:
        ws.title = title
    else:
        ws = wb.create_sheet(title)
    FillExcelWorksheetWithPropertyValuesOfElements(ws, propertyIds, elements)
    i += 1

excelFilePath = os.path.join(outputFolder, outputFileName)
wb.save(excelFilePath)
acu.OpenFile(excelFilePath)

if os.path.exists(excelFilePath):
    print("Saved Excel") #If you wanna know more join to our next Archicad Python AEC Workshop > https://pazacademy.ch/product/y2-python-archicad/?event=52
